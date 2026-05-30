"""
#63 — Exportación para contador (formato AFIP)

Genera un Excel con dos hojas:
  - Resumen: totales por categoría con IVA estimado
  - Detalle: cada transacción con fecha, descripción, categoría, cuenta, monto
"""
import io
from datetime import date
from decimal import Decimal
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import logging

logger = logging.getLogger(__name__)

IVA_RATE = Decimal("0.21")


def generate_tax_report_xlsx(
    transactions: List[Dict[str, Any]],
    period_start: date,
    period_end: date,
    user_name: str = "Usuario",
    deductible_category_ids: Optional[List[str]] = None,
) -> bytes:
    """
    Genera un archivo Excel para el contador.

    Args:
        transactions: lista de dicts con keys:
            id, description, transaction_date, amount, category_name,
            category_id, account_name
        period_start, period_end: rango del reporte
        user_name: nombre del usuario
        deductible_category_ids: si se pasan, solo incluye esas categorías

    Returns:
        bytes del archivo .xlsx
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl no está instalado. Ejecutar: pip install openpyxl")

    # Filtrar transacciones
    filtered = []
    for tx in transactions:
        tx_date = tx.get("transaction_date")
        if isinstance(tx_date, str):
            tx_date = date.fromisoformat(tx_date)

        if tx_date < period_start or tx_date > period_end:
            continue

        if deductible_category_ids and tx.get("category_id") not in deductible_category_ids:
            continue

        filtered.append({**tx, "transaction_date": tx_date})

    # Solo gastos (negativos)
    expenses = [tx for tx in filtered if float(tx.get("amount", 0)) < 0]
    expenses.sort(key=lambda x: x["transaction_date"])

    wb = Workbook()

    # ─── Hoja Resumen ─────────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Título
    ws_resumen.merge_cells("A1:E1")
    cell = ws_resumen["A1"]
    cell.value = f"Reporte Fiscal — {user_name}"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

    ws_resumen["A2"] = f"Período: {period_start.strftime('%d/%m/%Y')} - {period_end.strftime('%d/%m/%Y')}"
    ws_resumen["A2"].font = Font(italic=True, size=10)
    ws_resumen.merge_cells("A2:E2")

    # Headers de resumen
    row = 4
    headers = ["Categoría", "Cant. Movimientos", "Total Neto", "IVA Estimado (21%)", "Total con IVA"]
    for col, h in enumerate(headers, 1):
        c = ws_resumen.cell(row=row, column=col, value=h)
        c.font = subheader_font
        c.fill = subheader_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")

    # Agrupar por categoría
    by_category: Dict[str, Dict] = {}
    for tx in expenses:
        cat = tx.get("category_name", "Sin categoría")
        if cat not in by_category:
            by_category[cat] = {"count": 0, "total": Decimal("0")}
        by_category[cat]["count"] += 1
        by_category[cat]["total"] += abs(Decimal(str(tx["amount"])))

    row = 5
    grand_total = Decimal("0")
    grand_iva = Decimal("0")
    for cat_name, data in sorted(by_category.items()):
        neto = data["total"]
        iva = neto * IVA_RATE
        total_con_iva = neto + iva
        grand_total += neto
        grand_iva += iva

        ws_resumen.cell(row=row, column=1, value=cat_name).border = border
        ws_resumen.cell(row=row, column=2, value=data["count"]).border = border
        c = ws_resumen.cell(row=row, column=3, value=float(neto))
        c.number_format = '#,##0.00'
        c.border = border
        c = ws_resumen.cell(row=row, column=4, value=float(iva))
        c.number_format = '#,##0.00'
        c.border = border
        c = ws_resumen.cell(row=row, column=5, value=float(total_con_iva))
        c.number_format = '#,##0.00'
        c.border = border
        row += 1

    # Totales
    row += 1
    ws_resumen.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_resumen.cell(row=row, column=2, value=len(expenses)).font = Font(bold=True)
    c = ws_resumen.cell(row=row, column=3, value=float(grand_total))
    c.font = Font(bold=True)
    c.number_format = '#,##0.00'
    c = ws_resumen.cell(row=row, column=4, value=float(grand_iva))
    c.font = Font(bold=True)
    c.number_format = '#,##0.00'
    c = ws_resumen.cell(row=row, column=5, value=float(grand_total + grand_iva))
    c.font = Font(bold=True)
    c.number_format = '#,##0.00'

    # Ajustar anchos
    ws_resumen.column_dimensions["A"].width = 25
    ws_resumen.column_dimensions["B"].width = 18
    ws_resumen.column_dimensions["C"].width = 18
    ws_resumen.column_dimensions["D"].width = 22
    ws_resumen.column_dimensions["E"].width = 18

    # ─── Hoja Detalle ─────────────────────────────────────────
    ws_detalle = wb.create_sheet("Detalle")

    detail_headers = ["Fecha", "Descripción", "Categoría", "Cuenta", "Monto", "IVA Est.", "Total c/IVA"]
    for col, h in enumerate(detail_headers, 1):
        c = ws_detalle.cell(row=1, column=col, value=h)
        c.font = subheader_font
        c.fill = subheader_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")

    for i, tx in enumerate(expenses, 2):
        tx_date = tx["transaction_date"]
        monto = abs(Decimal(str(tx["amount"])))
        iva = monto * IVA_RATE

        ws_detalle.cell(row=i, column=1, value=tx_date.strftime("%d/%m/%Y")).border = border
        ws_detalle.cell(row=i, column=2, value=tx.get("description", "")).border = border
        ws_detalle.cell(row=i, column=3, value=tx.get("category_name", "")).border = border
        ws_detalle.cell(row=i, column=4, value=tx.get("account_name", "")).border = border
        c = ws_detalle.cell(row=i, column=5, value=float(monto))
        c.number_format = '#,##0.00'
        c.border = border
        c = ws_detalle.cell(row=i, column=6, value=float(iva))
        c.number_format = '#,##0.00'
        c.border = border
        c = ws_detalle.cell(row=i, column=7, value=float(monto + iva))
        c.number_format = '#,##0.00'
        c.border = border

    ws_detalle.column_dimensions["A"].width = 14
    ws_detalle.column_dimensions["B"].width = 35
    ws_detalle.column_dimensions["C"].width = 20
    ws_detalle.column_dimensions["D"].width = 20
    ws_detalle.column_dimensions["E"].width = 15
    ws_detalle.column_dimensions["F"].width = 15
    ws_detalle.column_dimensions["G"].width = 15

    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
