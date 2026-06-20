"""
Generación de reporte para contador (#63) — formato AFIP-friendly.

Toma las transacciones del período en las categorías marcadas como deducibles
y devuelve un PDF (resumen) y un Excel (detallado). El detalle Excel deja
columnas vacías para CUIT proveedor y discriminación de IVA — campos que el
contador completa luego.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from sqlalchemy.orm import Session, joinedload

from app.database import models


@dataclass
class ReportRow:
    transaction_date: date
    description: str
    category: str
    amount: Decimal


def collect_deductible_transactions(
    db: Session, user_id: str, start: date, end: date,
    category_ids: Optional[List[str]] = None,
) -> List[ReportRow]:
    """Trae todas las tx de gasto en categorías deducibles del user en el rango.

    Si `category_ids` está dado, restringe a esas (deben ser deducibles igualmente).
    """
    cat_filter = (models.Category.tax_deductible == True,)  # noqa: E712
    if category_ids:
        cat_filter = (
            models.Category.tax_deductible == True,  # noqa: E712
            models.Category.id.in_(category_ids),
        )
    rows = (
        db.query(models.Transaction)
        .join(models.Account)
        .join(models.Category)
        .filter(
            models.Account.user_id == user_id,
            models.Transaction.amount < 0,
            models.Transaction.transaction_date >= start,
            models.Transaction.transaction_date <= end,
            *cat_filter,
        )
        .options(joinedload(models.Transaction.category))
        .order_by(models.Transaction.transaction_date.asc())
        .all()
    )
    return [
        ReportRow(
            transaction_date=r.transaction_date,
            description=r.description,
            category=r.category.name if r.category else "—",
            amount=abs(Decimal(str(r.amount))),
        )
        for r in rows
    ]


def render_pdf_summary(user_name: str, start: date, end: date, rows: List[ReportRow]) -> bytes:
    """PDF tipo resumen para entregarle al contador: agrupado por categoría + total."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title="Reporte para contador — Vueltito",
    )
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("<b>Reporte de gastos deducibles</b>", styles["Title"]))
    story.append(Paragraph(f"Generado para: <b>{user_name}</b>", styles["Normal"]))
    story.append(Paragraph(
        f"Período: <b>{start.strftime('%d/%m/%Y')}</b> a <b>{end.strftime('%d/%m/%Y')}</b>",
        styles["Normal"],
    ))
    story.append(Spacer(1, 0.5 * cm))

    # Agrupación por categoría
    by_cat: dict[str, Decimal] = {}
    for r in rows:
        by_cat[r.category] = by_cat.get(r.category, Decimal("0")) + r.amount
    total = sum(by_cat.values(), Decimal("0"))

    cat_table_data = [["Categoría", "Total (ARS)"]]
    for cat, amount in sorted(by_cat.items(), key=lambda x: x[1], reverse=True):
        cat_table_data.append([cat, f"$ {amount:,.2f}"])
    cat_table_data.append(["TOTAL", f"$ {total:,.2f}"])

    cat_table = Table(cat_table_data, colWidths=[10 * cm, 5 * cm])
    cat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e5e7eb")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
    ]))
    story.append(cat_table)
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph(
        f"<b>{len(rows)}</b> transacciones consideradas.",
        styles["Normal"],
    ))
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph(
        "Detalle por transacción en archivo Excel adjunto. Las columnas CUIT y "
        "discriminación de IVA quedan vacías para que el contador las complete.",
        styles["Normal"],
    ))

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    return pdf


def render_excel_detail(user_name: str, start: date, end: date, rows: List[ReportRow]) -> bytes:
    """Excel con cada tx — incluye columnas CUIT / IVA / Neto para que el contador complete."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)

    ws["A1"] = f"Reporte de gastos deducibles — {user_name}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    ws["A2"] = f"Período: {start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}"
    ws.merge_cells("A2:H2")

    headers = ["Fecha", "Descripción", "Categoría", "Total (ARS)", "CUIT proveedor", "Neto", "IVA 21%", "Otros"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 5
    for r in rows:
        ws.cell(row=row_idx, column=1, value=r.transaction_date.strftime("%d/%m/%Y"))
        ws.cell(row=row_idx, column=2, value=r.description)
        ws.cell(row=row_idx, column=3, value=r.category)
        ws.cell(row=row_idx, column=4, value=float(r.amount))
        # cols 5, 6, 7, 8 quedan vacías para el contador
        row_idx += 1

    total = sum((r.amount for r in rows), Decimal("0"))
    ws.cell(row=row_idx, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_idx, column=4, value=float(total)).font = Font(bold=True)

    # Ajuste de ancho aproximado
    widths = [12, 35, 18, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        col_letter = ws.cell(row=4, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data
